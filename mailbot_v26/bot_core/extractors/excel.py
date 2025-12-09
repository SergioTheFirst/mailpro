"""Excel text extraction using openpyxl."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def extract_excel_text(excel_bytes: bytes) -> str:
    """
    Extract text from XLS/XLSX.

    Args:
        excel_bytes: Raw Excel file content

    Returns:
        Extracted text (all cells concatenated) or empty string
    """
    try:
        wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

        text_parts = []
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)

        wb.close()
        return "\n".join(text_parts).strip()

    except Exception:
        return ""


# Self-test
if __name__ == "__main__":
    # Similar to DOCX, hard to create Excel without library
    # Verify error handling
    try:
        result = extract_excel_text(b"invalid excel")
        assert result == ""
        print("✅ Excel extractor self-test PASSED (error handling verified)")
    except Exception:
        print("❌ Excel extractor failed unexpectedly")
